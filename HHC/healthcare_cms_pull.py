#!/usr/bin/env python3
import io
import os
import json
import logging
import zipfile
import argparse
import multiprocessing
import concurrent.futures
import urllib.request as request

import openpyxl

from json_list_parser import json_list_parser
import models
import db

NULL_URL = "NOT SUBMITTED"
LOGGER = logging.getLogger('HHC')
LOGGER.addHandler(logging.FileHandler("healthcare_cms_pull.log"))

CONFIG = {}


def get_cms_spreadsheet(url):
    LOGGER.info("Pulling CMS Spreadsheet from {}".format(url))
    with request.urlopen(url) as f:
        b = io.BytesIO(f.read())
    zf = zipfile.ZipFile(b)
    fname = zf.namelist()[0]
    nb = openpyxl.load_workbook(zf.open(fname))
    LOGGER.info("Sucessfully pulled CMS Spreadsheet".format(url))
    return nb.active


def get_issuers_from_cms_spreadsheet(cms_url):
    ws = get_cms_spreadsheet(cms_url)
    issuers = []
    current_row = 1
    LOGGER.info("BEGIN PARSING CMS SPREADSHEET")
    while ws.cell(row=current_row+1, column=1).value:
        current_row += 1
        i = models.Issuer()
        i.id_issuer            = ws.cell(row=current_row, column=3).value
        i.name                 = ws.cell(row=current_row, column=4).value
        i.marketplace_category = ws.cell(row=current_row, column=1).value
        i.url_submitted        = ws.cell(row=current_row, column=5).value
        i.state                = ws.cell(row=current_row, column=2).value.lower()
        i.plans = []
        if i.url_submitted == NULL_URL:
            LOGGER.warning("Missing json url for Issuer: {}, {}".format(i.name, i.id_issuer))
        LOGGER.info("Issuer Parsed: {}".format(i.name))
        issuers.append(i)
    LOGGER.info("FINISH PARSING CMS SPREADSHEET")
    return issuers


def pull_issuer_index(issuer):
    if issuer.url_submitted == NULL_URL:
        return (issuer, [], [])
    LOGGER.info("PULLING JSON INDEX FOR ISSUER {} @ {}".format(issuer.name,
                                                                issuer.url_submitted))
    plan_urls = []
    provider_urls = []
    try:
        with request.urlopen(issuer.url_submitted, timeout=20) as conn:
            js = json.loads(conn.read().decode('utf8'))
            plan_urls = js['plan_urls']
            provider_urls = js['provider_urls']
            LOGGER.info("FINISHED PULLING JSON INDEX FOR {}".format(issuer.name))
    except Exception as e:
        LOGGER.error("ERROR LOADING JSON INDEX FOR {}, {}".format(issuer.name, e))
    return (issuer, plan_urls, provider_urls)


def pull_plan_fulldata(conn, issuer, plan_url):
    LOGGER.info("Pulling plan page at {}".format(plan_url))
    plans = {}
    try:
        req = request.urlopen(plan_url, timeout=20)
        file_size = req.info().get('Content-Length')
        try:
            file_size = int(file_size)
        except ValueError:
            file_size = 0
        LOGGER.info("Downloading plan json\n\tFile Size: {}Bytes".format(file_size))
        json_objs = json_list_parser(req, file_size)
        for i,((bytes_read,file_size),plan_dict) in enumerate(json_objs):
            try:
                plan = models.build_plan_from_dict(issuer, plan_dict)
                if plan.id_plan in plans:
                    LOGGER.warning("duplicate plan id: {}".format(plan.id_plan))
                plans[plan.id_plan] = plan
            except Exception as e:
                LOGGER.exception(e)
                LOGGER.info(plan_dict)
                continue
        db.insert_plans(conn, plans.values())
        conn.commit()
    except Exception as e:
        LOGGER.exception(e)
        LOGGER.warning("Error loading plan data from {}".format(issuer.name))
    return plans


def pull_provider_fulldata(conn, issuer, plans, provider_url, config):
    try:
        req = request.urlopen(provider_url)
        json_objs = json_list_parser(req)
        providers = []
        for i,((bytes_read,file_size),provider_dict) in enumerate(json_objs):
            try:
                provider, fake_plans = models.build_provider_from_dict(issuer, plans, provider_dict, config)
                if fake_plans:
                    LOGGER.warning("Found provider {} with undocumented plans: {}".format(provider.name,fake_plans.keys()))
                db.insert_plans(conn, fake_plans.values())
                plans.update(fake_plans)
                providers.append(provider)
            except Exception as e:
                LOGGER.exception(e)
                LOGGER.info(str(provider_dict))
                continue
            if((i%1000) == 0):
                if file_size > 0:
                    progress = int(100*(bytes_read/file_size))
                    LOGGER.info("Downloaded ({}/{}) {:02d}%".format(bytes_read,file_size,progress))
                else:
                    LOGGER.info("Downloaded ({}/{})".format(bytes_read,file_size))
                LOGGER.info("Parsed {} Providers".format(i))
                db.insert_providers(conn, providers)
                conn.commit()
                providers.clear()
        db.insert_providers(conn, providers)
        conn.commit()
    except Exception as e:
        LOGGER.exception(e)
        LOGGER.warning("Error loading plan plan data from {}".format(issuer.name))


def init_full_pull_logger(id_issuer):
    global LOGGER
    LOGGER = logging.Logger("FULLPULL:"+str(id_issuer))
    log_name = os.path.join("db","{}.log".format(id_issuer))
    LOGGER.addHandler(logging.FileHandler(log_name))


def pull_issuer_fulldata(args):
    id_issuer, requested_states, config = args
    init_full_pull_logger(id_issuer)
    conn = db.open_db(id_issuer)
    issuer = db.query_issuer(conn, id_issuer)
    if requested_states is not None and issuer.state not in requested_states:
        db.close_db(conn)
        return
    LOGGER.info("PULLING FULLDATA FOR ISSUER {}".format(id_issuer))
    plan_urls, provider_urls = db.query_issuer_urls(conn, issuer)
    plans = {}
    LOGGER.info("BEGIN PULLING PLAN DATA")
    for plan_url in plan_urls:
        plans.update(pull_plan_fulldata(conn, issuer, plan_url))
    LOGGER.info("PULLED {} PLANS FOR ISSUER {}".format(len(plans),id_issuer))

    LOGGER.info("BEGIN PULLING PROVIDER DATA")
    n = len(provider_urls)
    for i, provider_url in enumerate(provider_urls):
        LOGGER.info("Pulling provider page {} of {} at {}".format(i+1, n, provider_url))
        pull_provider_fulldata(conn, issuer, plans, provider_url, config)
    db.close_db(conn)
    LOGGER.info("FINISHED PULLING PROVIDER DATA")

    LOGGER.info("FINISHED PULLING FULLDATA FOR ISSUER: {}, {}".format(id_issuer, issuer.name))


def pull_issuers_fulldata(requested_ids, requested_states, processes, config):
    db_filenames = [os.path.splitext(fname) for fname in os.listdir("db")]
    ids_issuer = [int(base) for (base,ext) in db_filenames if ext == ".sqlite3"]
    if requested_ids is not None:
        ids_issuer = [id_issuer for id_issuer in ids_issuer if id_issuer in requested_ids]
    args = [(id_issuer, requested_states, config) for id_issuer in ids_issuer]
    with multiprocessing.Pool(processes) as p:
        p.map(pull_issuer_fulldata, args)


def pull_issuers_metadata(cms_url, requested_ids, requested_states, config):
    issuers = get_issuers_from_cms_spreadsheet(cms_url)
    if requested_ids is not None:
        issuers = [issuer for issuer in issuers if issuer.id_issuer in requested_ids]
    if requested_states is not None:
        issuers = [issuer for issuer in issuers if issuer.state.lower() in requested_states]
    if not os.path.isdir("db"):
        os.mkdir("db")

    with concurrent.futures.ThreadPoolExecutor(max_workers=30) as executor:
        futures = [executor.submit(pull_issuer_index, issuer) for issuer in issuers]
        for future in concurrent.futures.as_completed(futures):
            issuer, plan_urls, provider_urls = future.result()
            conn = db.init_db(issuer.id_issuer)
            db.insert_issuer(conn, issuer)
            db.insert_issuer_urls(conn, issuer, plan_urls, provider_urls)
            db.close_db(conn)


def main():
    parser = argparse.ArgumentParser(description="Utility to Download Insurance Acceptance Data")
    add = parser.add_argument
    add('--cmsurl', default="http://download.cms.gov/marketplace-puf/2016/machine-readable-url-puf.zip",
            help = "the url to \"machine-readable-url-puf.zip\"")
    add('--fulladdress', action='store_true',
            help = "Enable to store full address data, leave out to conserve disk space")
    add('--issuer_ids', default = None, nargs='+', type=int,
            help="Specify specific issuers to do fulldata pull on")
    add('--states', default = None, nargs='+', type=str,
            help="Specify a list of specific states to download fulldata on")
    add('--processes', default = 1, type=int,
            help="Set the number of processes to use in the full data pull")
    add('action', choices=['init','download', 'all'],
            help="Action to undertake")
    args = parser.parse_args()

    config = {}
    config['FULL_ADDRESS'] = args.fulladdress
    if args.action in ('init', 'all'):
        pull_issuers_metadata(args.cmsurl, args.issuer_ids, args.states, config)
    if args.action in ('download', 'all'):
        pull_issuers_fulldata(args.issuer_ids, args.states, args.processes, config)


if __name__ == "__main__":
    main()
