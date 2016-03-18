#!/usr/bin/env python3
import io
import os
import json
import logging
import zipfile
import argparse
import collections
import multiprocessing as mp
import urllib.parse as parse
import urllib.request as request

import openpyxl

from json_list_parser import json_list_parser, download_formatter
import models
import db

NULL_URL = "NOT SUBMITTED"
PROFILE_DB = False
PROFILE_DOWNLOAD = False


def init_logger(logger_name):
    if not os.path.exists("data"):
        os.mkdir("data")
    logger = logging.Logger(logger_name)
    log_name = "data/{}.log".format(logger_name)
    ch = logging.FileHandler(log_name, mode='w')
    fmt = "%(asctime)s [%(name)s] %(levelname)s: %(message)s"
    ch.setFormatter(logging.Formatter(fmt=fmt))
    logger.addHandler(ch)
    logger.setLevel(logging.INFO)
    return logger


class Downloader:
    url_limit = 0
    data_limit = 0
    download_attempts = 3

    def __init__(self, issuer_group, queue, label):
        self.issuer_group = issuer_group
        self.issuer_group.idx_issuer_group = label
        for issuer in self.issuer_group.issuers:
            issuer.idx_issuer_group = label
        self.q = queue
        self.logger = init_logger("DL_{}".format(label))

    def run(self):
        log = self.logger
        success = self._download_index()
        status = "finished" if success else "failed"
        self.issuer_group.index_status = status
        self.q.put(self.issuer_group)
        for issuer in self.issuer_group.issuers:
            self.q.put(issuer)
        if not success:
            return

        plan_urls = list(filter(lambda x: x.url_type == models.URLType.plan,
                                self.issuer_group.data_urls))
        n = len(plan_urls)
        for i, url in enumerate(plan_urls):
            log.info("Downloading from Plan URL {}/{}".format(i+1, n))
            success = self._download_objects(url.url, models.Plan, 0)
            url.status = "finished" if success else "failed"
            self.q.put(url)

        data_urls = list(filter(lambda x: x.url_type != models.URLType.plan,
                                self.issuer_group.data_urls))
        if self.url_limit:
            data_urls = data_urls[:self.url_limit]
        n = len(data_urls)
        for i, url in enumerate(data_urls):
            log.info("Downloading from Data URL {}/{}".format(i+1, n))
            if url.url_type == models.URLType.drug:
                success = self._download_objects(url.url, models.Drug)
            else:  # provider url
                success = self._download_objects(url.url, models.Provider)
            url.status = "finished" if success else "failed"
            self.q.put(url)

    def _download_index(self):
        log = self.logger
        iss_grp = self.issuer_group
        idx = iss_grp.idx_issuer_group
        if iss_grp.index_url == NULL_URL:
            return
        log.info("Pulling JSON index from \"{}\"".format(iss_grp.index_url))
        try:
            with request.urlopen(iss_grp.index_url, timeout=20) as conn:
                js = json.loads(conn.read().decode('utf8'))
                plans = [models.IssuerGroupURL(idx, url, models.URLType.plan)
                         for url in js['plan_urls']]
                provs = [models.IssuerGroupURL(idx, url, models.URLType.prov)
                         for url in js['provider_urls']]
                drugs = [models.IssuerGroupURL(idx, url, models.URLType.drug)
                         for url in js['formulary_urls']]
                iss_grp.data_urls = (plans+provs+drugs)

                fmt = "Finished pulling JSON index from URL \"{}\""
                log.info(fmt.format(iss_grp.index_url))
                return True
        except Exception as e:
            log.exception(e)
            fmt = "Failed to load JSON index from URL \"{}\""
            log.error(fmt.format(iss_grp.index_url))
            return False

    def _download_objects_attempt(self, url, class_, data_limit=None):
        log = self.logger
        if data_limit is None:
            data_limit = self.data_limit
        formatter = download_formatter()

        def status(bytes_read, file_size, i):
            s = formatter(bytes_read, file_size)
            log.info("Downloaded {}".format(s))
            log.info("Parsed {} data objects".format(i))
        try:
            json_objs = json_list_parser(url)
            for i, ((bytes_read, file_size), obj_dict) in enumerate(json_objs):
                try:
                    obj = class_(obj_dict)
                    self.q.put(obj)
                except Exception as e:
                    log.exception(e)
                    log.info(obj_dict)
                    continue
                if i > 0 and ((i % 1000) == 0):
                    status(bytes_read, file_size, i)
                if data_limit and i > data_limit:
                    log.info("Hit data limit, breaking...")
                    status(bytes_read, file_size, i)
                    break
            fmt = "Finished download of url: {} |{} MB"
            log.info(fmt.format(url, bytes_read/2**20))
            return True
        except Exception as e:
            log.exception(e)
            log.warning("Error loading data from {}".format(url))
            return False

    def _download_objects(self, url, class_, data_limit=None):
        for i in range(self.download_attempts):
            fmt = "Starting Download Attempt {}/{}"
            self.logger.info(fmt.format(i+1, self.download_attempts))
            success = self._download_objects_attempt(url, class_,
                                                     data_limit=None)
            if success:
                return True
            else:
                continue
        return False


class Consumer:
    commit_size = 10000

    def __init__(self, queue, label="CS", states=None):
        self.logger = init_logger("DL_{}".format(label))
        self.q = queue
        self.conn = db.init_db()
        self.states = states

        # Declare a few auxillary lookup tables
        self.provs = {}  # maps npi to idx_provider
        self.drugs = {}  # maps rx_norm_id to idx_drug
        self.facility_types = {}  # maps name(str) to idx
        self.specialties = {}  # maps name(str) to idx
        self.languages = {}  # maps name(str) to idx
        self.plans = {}  # maps (id_issuer,id_plan) to idx_plan
        self.commit_obj_cnt = 0

    def _process_issuer_group(self, issuer_group):
        id_ = issuer_group.idx_issuer_group
        self.logger.debug("Inserting Issuer Group: {}".format(id_))
        db.insert_issuer_group(self.conn, issuer_group)

    def _process_url(self, url):
        self.logger.debug("Inserting URL: {}".format(url))
        db.insert_data_url(self.conn, url)

    def _process_issuer(self, issuer):
        self.logger.debug("Inserting Issuer: {}".format(issuer.id_issuer))
        db.insert_issuer(self.conn, issuer)

    def _process_plan(self, plan):
        self.logger.debug("Inserting Plan: {}".format(plan.marketing_name))
        if (plan.id_issuer, plan.id_plan) not in self.plans:
            idx_plan = db.insert_plan(self.conn, plan)
            self.plans[(plan.id_issuer, plan.id_plan)] = idx_plan

    def _check_provider_in_state(self, prov):
        if not self.states:
            return True  # No states selected
        for addr in prov.addresses:
            if addr.state in self.states:
                return True
        return False

    def _process_provider(self, prov):
        conn = self.conn
        log = self.logger
        log.debug("Inserting Provider: {},{}".format(prov.npi, prov.name))
        if not prov.npi:
            return
        if not self._check_provider_in_state(prov):
            return
        if prov.npi not in self.provs:
            self.provs[prov.npi] = db.insert_provider(conn, prov)
            idx_prov = self.provs[prov.npi]

            for lang in prov.languages:
                if lang not in self.languages:
                    self.languages[lang] = db.insert_language(conn, lang)
                idx_lang = self.languages[lang]
                db.insert_provider_language(conn, idx_prov, idx_lang)

            for spec in prov.specialties:
                if spec not in self.specialties:
                    self.specialties[spec] = db.insert_specialty(conn, spec)
                idx_spec = self.specialties[spec]
                db.insert_provider_specialty(conn, idx_prov, idx_spec)

            for ft in prov.facility_types:
                if ft not in self.facility_types:
                    self.facility_types[ft] = db.insert_facility_type(conn, ft)
                idx_facil = self.facility_types[ft]
                db.insert_provider_facility_type(conn, idx_prov, idx_facil)

            for address in prov.addresses:
                db.insert_address(conn, address, idx_prov)
        idx_prov = self.provs[prov.npi]

        for plan in prov.plans:
            if (plan.id_issuer, plan.id_plan) not in self.plans:
                p = models.Plan()
                p.id_issuer = plan.id_issuer
                p.id_plan = plan.id_plan
                p.plan_id_type = "VOID"
                self._process_plan(p)
            idx_plan = self.plans[(plan.id_issuer, plan.id_plan)]
            db.insert_provider_plan(conn, plan, idx_plan, idx_prov)

    def _process_drug(self, drug):
        conn = self.conn
        log = self.logger
        log.debug("Inserting Drug: {}".format(drug.name))
        if not drug.rxnorm_id:
            return
        if drug.rxnorm_id not in self.drugs:
            self.drugs[drug.rxnorm_id] = db.insert_drug(conn, drug)
        idx_drug = self.drugs[drug.rxnorm_id]
        for plan in drug.plans:
            db.insert_drug_plan(conn, plan, idx_drug)

    def run(self):
        log = self.logger
        map_ = {models.IssuerGroup:      self._process_issuer_group,
                models.IssuerGroupURL:   self._process_url,
                models.Issuer:           self._process_issuer,
                models.Provider:         self._process_provider,
                models.Plan:             self._process_plan,
                models.Drug:             self._process_drug}
        while True:
            obj = self.q.get()
            if obj == "QUIT":
                log.info("Consumer received quit signal. closing db...")
                self.conn.commit()
                self.conn.close()
                break
            else:
                try:
                    map_[type(obj)](obj)
                    self.commit_obj_cnt += 1
                    if self.commit_obj_cnt >= self.commit_size:
                        fmt = ("Successfully downloaded {} objects, "
                               "commiting to db.")
                        log.info(fmt.format(self.commit_obj_cnt))
                        self.conn.commit()
                        log.info("Finished commit.")
                        self.commit_obj_cnt = 0
                except Exception as e:
                    log.exception(e)


def consume(q, states):
    consumer = Consumer(q, states=states)
    if PROFILE_DB:
        import cProfile
        cProfile.runctx('consumer.run()', globals(), locals(),
                        'consumer.prof')
    else:
        consumer.run()


def produce(args):
    issuer_group, label = args
    producer = Downloader(issuer_group, produce.q, label)
    if PROFILE_DOWNLOAD:
        import cProfile
        fname = "producer_{:02d}.prof".format(label)
        cProfile.runctx('producer.run()', globals(), locals(), fname)
    else:
        producer.run()


def init_produce(q):
    produce.q = q


class Manager:
    def __init__(self, cms_url, filters, num_processes=10):
        self.logger = init_logger("MANAGER")
        self.cms_url = cms_url
        self.requested_issuer_ids = filters['issuer_ids']
        self.requested_states = filters['states']
        self.num_processes = num_processes

    def _apply_filters(self):
        log = self.logger
        ids = self.requested_issuer_ids
        states = self.requested_states
        if ids:
            log.info("Filtering on ids: {}".format(ids))
        else:
            log.info("No ids specified, keeping all.")
        if states:
            log.info("Filtering on states: {}".format(states))
        else:
            log.info("No states specified, keeping all.")

        def filt(issuers):
            return [i for i in issuers
                    if ((not states or i.state in states) and
                        (not ids or i.id_issuer in ids))]
        grp_filt = [group
                    for group in self.issuer_groups if filt(group.issuers)]
        log.info("After filtering, will download the following groups:")
        for i, grp in enumerate(grp_filt):
            log.info("Group #{}, index-url: \"{}\"".format(i, grp.index_url))
            iss_ids = [iss.id_issuer for iss in grp.issuers]
            log.info("\tWith Issuers: {}".format(iss_ids))
        self.issuer_groups = grp_filt

    def _find_issuer_groups(self):
        log = self.logger
        parse_result = parse.urlparse(self.cms_url)
        if parse_result.scheme == "file":
            fmt = "Opening local CMS Spreadsheet from {}"
            log.info(fmt.format(parse_result.path))
            nb = openpyxl.load_workbook(open(parse_result.path, 'rb'))
        else:  # Web URL
            log.info("Pulling CMS Spreadsheet from {}".format(self.cms_url))
            with request.urlopen(self.cms_url) as f:
                b = io.BytesIO(f.read())
            zf = zipfile.ZipFile(b)
            fname = zf.namelist()[0]
            nb = openpyxl.load_workbook(zf.open(fname))
            log.info("Sucessfully pulled CMS Spreadsheet".format(self.cms_url))

        issuer_groups = collections.OrderedDict()
        current_row = 1
        log.info("BEGIN PARSING CMS SPREADSHEET")
        ws = nb.active
        while ws.cell(row=current_row+1, column=1).value:
            current_row += 1
            i = models.Issuer()
            i.state = ws.cell(row=current_row, column=1).value.lower()
            i.id_issuer = ws.cell(row=current_row, column=2).value
            i.name = ws.cell(row=current_row, column=3).value
            index_url = ws.cell(row=current_row, column=4).value
            if index_url == NULL_URL:
                fmt = "Missing JSON url for Issuer: {}, {}"
                log.warning(fmt.format(i.name, i.id_issuer))
            log.info("Issuer Parsed: {}".format(i.name))
            if index_url not in issuer_groups:
                issuer_groups[index_url] = []
            issuer_groups[index_url].append(i)
        log.info("FINISH PARSING CMS SPREADSHEET")

        group_objs = []
        for url, issuers in issuer_groups.items():
            issuer_group = models.IssuerGroup()
            issuer_group.index_url = url
            issuer_group.issuers = issuers
            group_objs.append(issuer_group)
        self.issuer_groups = group_objs

    def run(self):
        self._find_issuer_groups()
        self._apply_filters()
        q = mp.Queue(maxsize=1000)
        consume_proc = mp.Process(target=consume, args=(q,
                                                        self.requested_states))
        consume_proc.start()
        pool = mp.Pool(self.num_processes, init_produce, [q])
        pool.map(produce, [(grp, i)
                           for i, grp in enumerate(self.issuer_groups)])
        q.put("QUIT")
        consume_proc.join()


def main():
    desc = "Utility to Download Insurance Acceptance Data"
    parser = argparse.ArgumentParser(description=desc)
    add = parser.add_argument
    urlhelp = ("the url to \"machine-readable-url-puf.zip\", or the path "
               "to \"Machine_Readable_PUF_*.xlsx\" in the form "
               "\"file:/path/to/file\"")
    urldefault = ("http://download.cms.gov/marketplace-puf/2016/"
                  "machine-readable-url-puf.zip")
    add('--cmsurl', default=urldefault, help=urlhelp)
    add('--issuerids', default=None, nargs='+', type=int,
        help="Specify specific issuers to do fulldata pull on")
    add('--states', default=None, nargs='+', type=str,
        help="Specify a list of specific states to download fulldata on")
    add('--processes', default=1, type=int,
        help="Set the number of processes to use in the full data pull")
    args = parser.parse_args()

    filters = {'issuer_ids': args.issuerids,
               'states': args.states}
    manager = Manager(args.cmsurl, filters, args.processes)
    manager.run()


if __name__ == "__main__":
    main()
