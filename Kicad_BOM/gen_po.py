#!/usr/bin/env python3
from collections import namedtuple, defaultdict
from functools import lru_cache, wraps
from os.path import split, join
from decimal import Decimal
import argparse
import urllib
import re
import os
import json
import datetime

import openpyxl
import requests

BLANK_REQ = "http://www.unl.edu/physics/docs/Requisition2014.xlsx"
BLANK_REQ_FILE = "/tmp/empty_form.xlsx"
ARGS = None


def chunks(l, n):
    """Yield successive n-sized chunks from l"""
    for i in range(0, len(l), n):
        yield l[i:i+n]


def fetch_empty_form():
    form_data = requests.get(BLANK_REQ).content
    with open(BLANK_REQ_FILE, "wb") as f:
        f.write(form_data)


def delete_empty_form():
    os.remove(BLANK_REQ_FILE)


def get_part_list():
    with open(ARGS.json_in, 'r') as f:
        return json.load(f)


def populate_misc_fields(workbook):
    ws = workbook.active
    ws['B11'] = 'Digikey inc.'
    ws['B13'] = '701 Brooks Avenue South'
    ws['B16'] = 'Thief River Falls, MN 56701 USA'
    ws['B18'] = ARGS.usage
    ws['B22'] = '218-681-6674'
    ws['F22'] = '218-681-3380'
    ws['B26'] = 'www.digikey.com'
    ws['E28'] = ARGS.delivery_date
    ws['D43'] = ARGS.cost_object
    ws['B45'] = datetime.date.today().strftime("%b. %d, %Y")
    ws['C47'] = ARGS.requestor_name
    ws['K47'] = ARGS.requestor_phone
    ws['C49'] = ARGS.supervisor_name


def clear_parts(workbook):
    ws = workbook.active
    for i in range(32, 42):
        i = str(i)
        ws["A"+i] = ""
        ws["B"+i] = ""
        ws["L"+i] = ""
        ws["Q"+i] = ""


def populate_parts(workbook, parts):
    clear_parts(workbook)
    ws = workbook.active
    for i, part in enumerate(parts):
        i = str(i+32)
        ws["A"+i] = part['part_number']
        ws["B"+i] = part['description']
        ws["L"+i] = part['num_order']
        ws["Q"+i] = part['unit_price'][1:]


def place_sheet_number(workbook, i, n):
    ws = workbook.active
    ws["A52"] = "Sheet {} of {}".format(i, n)


def main():
    fetch_empty_form()
    parts = get_part_list()
    workbook = openpyxl.load_workbook(BLANK_REQ_FILE)
    populate_misc_fields(workbook)

    parts_chunks = list(chunks(parts, 10))
    for i, parts_chunk in enumerate(chunks(parts, 10)):
        i = i+1
        populate_parts(workbook, parts_chunk)
        place_sheet_number(workbook, i, len(parts_chunks))

        fname = "digikey_req{:02d}.xlsx".format(i)
        workbook.save(join(ARGS.output_dir, fname))


if __name__ == "__main__":
    PARSER = argparse.ArgumentParser()
    PARSER.add_argument("json_in",
                        help="json output from BOM page")
    PARSER.add_argument("--output_dir", dest='output_dir',
                        default="",
                        help="Directory to place resulting POs")
    PARSER.add_argument("--usage", dest='usage',
                        default="",
                        help="How is purchase used for research project?")
    PARSER.add_argument("--delivery_date", dest='delivery_date',
                        default="",
                        help="Required Delivery date")
    PARSER.add_argument("--cost_object", dest='cost_object',
                        default="",
                        help="Cost Object")
    PARSER.add_argument("--requestor_name", dest='requestor_name',
                        default="",
                        help="Requestor's Name")
    PARSER.add_argument("--requestor_phone", dest='requestor_phone',
                        default="",
                        help="Requestor's Phone Number")
    PARSER.add_argument("--supervisor_name", dest='supervisor_name',
                        default="",
                        help="Supervisor's Name")
    ARGS = PARSER.parse_args()
    main()

