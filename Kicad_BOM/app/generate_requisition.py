from io import BytesIO
import urllib
import zipfile
import datetime

import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

from .vendor_fetch import vendors

BLANK_REQ_URL = "http://www.unl.edu/physics/docs/Requisition2014.xlsx"
_BLANK_FORM_RAW = None


def chunks(l, n):
    """Yield successive n-sized chunks from l"""
    cs = []
    for i in range(0, len(l), n):
        cs.append(l[i:i+n])
    return cs


def fetch_empty_form():
    global _BLANK_FORM_RAW
    if _BLANK_FORM_RAW is None:
        _BLANK_FORM_RAW = urllib.request.urlopen(BLANK_REQ_URL).read()
    return openpyxl.load_workbook(BytesIO(_BLANK_FORM_RAW))


def populate_misc_fields(workbook, order, vendor):
    ws = workbook.active
    ws['B11'] = vendor.form_info['long_name']
    ws['B13'] = vendor.form_info['addr_line1']
    ws['B16'] = vendor.form_info['addr_line2']
    ws['B18'] = order.description
    ws['B22'] = vendor.form_info['phone']
    ws['F22'] = vendor.form_info['fax']
    ws['B26'] = vendor.form_info['website']
    ws['E28'] = str(order.delivery_date)
    ws['D43'] = order.cost_object
    ws['B45'] = datetime.date.today().strftime("%b. %d, %Y")
    ws['C47'] = order.requestor_name
    ws['K47'] = order.requestor_phone
    ws['C49'] = order.supervisor_name


def populate_parts(workbook, parts):
    ws = workbook.active
    for i, part in enumerate(parts):
        i = str(i+32)
        ws["A"+i] = part.vendorpart.vendor_part_number
        ws["B"+i] = part.vendorpart.part.short_description
        n = part.number_ordered
        ws["L"+i] = n
        try:
            ws["Q"+i] = part.vendorpart.get_pricebreak(n)
        except ValueError:
            ws["Q"+i] = "N/A"


def place_sheet_number(workbook, i, n):
    ws = workbook.active
    ws["A52"] = "Sheet {} of {}".format(i, n)


def build_requisition(order):
    parts_chunked = chunks(order.vendorparts, 10)
    bio = BytesIO()
    zf = zipfile.ZipFile(bio, 'x')
    for i, parts_chunk in enumerate(parts_chunked):
        workbook = fetch_empty_form()
        populate_parts(workbook, parts_chunk)
        populate_misc_fields(workbook, order, vendors['Digikey'])
        place_sheet_number(workbook, i+1, len(parts_chunked))

        fname = "digikey_req{:02d}.xlsx".format(i+1)
        zf.writestr(fname, save_virtual_workbook(workbook))
    zf.close()
    bio.seek(0)
    return bio.read()


if __name__ == '__main__':
    from .models import Order
    order = Order.query.all()[-1]
    with open('outfile.zip', 'wb') as f:
        f.write(build_requisition(order))
